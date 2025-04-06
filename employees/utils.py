import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime

def generate_excel(queryset):
    # ワークブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '従業員一覧'
    
    # ヘッダー行のスタイル
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # ヘッダー行の設定
    headers = ['社員ID', '氏名', '所属部署', '性別', '生年月日', '入社日', '在籍状況', 'メールアドレス', '給与']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # データ行の設定
    for row_num, employee in enumerate(queryset, 2):
        # 列の順番はヘッダーと一致させる
        ws.cell(row=row_num, column=1).value = employee.employee_id
        ws.cell(row=row_num, column=2).value = employee.name
        ws.cell(row=row_num, column=3).value = employee.department.name
        ws.cell(row=row_num, column=4).value = employee.get_gender_display()
        ws.cell(row=row_num, column=5).value = employee.date_of_birth
        ws.cell(row=row_num, column=5).number_format = 'yyyy/mm/dd'
        ws.cell(row=row_num, column=6).value = employee.hire_date
        ws.cell(row=row_num, column=6).number_format = 'yyyy/mm/dd'
        ws.cell(row=row_num, column=7).value = employee.get_status_display()
        ws.cell(row=row_num, column=8).value = employee.email
        ws.cell(row=row_num, column=9).value = employee.salary
        ws.cell(row=row_num, column=9).number_format = '#,##0'
    
    # 列幅の自動調整
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # HttpResponseオブジェクトを作成してExcelファイルを返す
    current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=employees_{current_datetime}.xlsx'
    
    # ExcelファイルをHttpResponseに書き込む
    wb.save(response)
    return response